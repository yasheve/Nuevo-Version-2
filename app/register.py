"""Luminaire register rendering — shared by the CSV and XLSX export endpoints.

ONE column spec drives BOTH outputs, so the CSV and the Excel can never drift
apart. Callers pass an iterable of Asset-like objects (anything exposing the
read attributes below); this module imports nothing from the rest of the app
(no DB, no storage, no settings) so it stays import-light and unit-testable.

Column set (24 columns) matches the agreed register layout:
  * No standalone `City` column; `Town` falls back to the legacy `city` value for
    pre-redesign rows so it is never blank for historical captures.
  * `Company` defaults to "eThekwini Municipality" for EMP (eThekwini) captures,
    which otherwise carry no company; EMC captures keep their contractor company.
"""
from datetime import datetime, timezone
import csv
import io

MUNICIPALITY = "eThekwini Municipality"


def _iso(dt):
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat().replace("+00:00", "Z")


def _full_name(a):
    return " ".join(p for p in (a.captured_by_name, a.captured_by_surname) if p)


def _company(a):
    # EMP (eThekwini) staff carry no contractor company -> default to the
    # municipality. EMC captures keep the company stamped from their token.
    if (a.captured_by_kind or "").upper() == "EMP":
        return a.company_name or MUNICIPALITY
    return a.company_name


# (header, kind, getter). kind drives per-format rendering:
#   text | int | float | datetime | imei (text, kept verbatim so long IMEIs
#   never collapse to scientific notation in Excel).
REGISTER_SPEC = [
    ("Luminaire Manufacturer",  "text",     lambda a: a.manufacturer),
    ("Luminaire Model",         "text",     lambda a: a.model_no),
    ("Luminaire Serial",        "text",     lambda a: a.serial_no),
    ("Luminaire Year",          "int",      lambda a: a.manufacture_year),
    ("Luminaire Wattage",       "int",      lambda a: a.wattage),
    ("Controller Manufacturer", "text",     lambda a: a.controller_manufacturer),
    ("Controller Model",        "text",     lambda a: a.controller_model),
    ("Controller IMEI",         "imei",     lambda a: a.imei),
    ("Date of Installation",    "datetime", lambda a: a.captured_at),
    ("Work Order Number",       "text",     lambda a: a.work_order_no),
    ("Installer Type",          "text",     lambda a: a.captured_by_kind),
    ("Installer Name",          "text",     _full_name),
    ("Designation",             "text",     lambda a: a.designation),
    ("Service Number",          "text",     lambda a: a.service_no),
    ("Company",                 "text",     _company),
    ("Contractor Number",       "text",     lambda a: a.contractor_number),
    ("Region",                  "text",     lambda a: a.region),
    ("Road Name",               "text",     lambda a: a.road),
    ("Suburb",                  "text",     lambda a: a.suburb),
    ("Town",                    "text",     lambda a: a.town or a.city),
    ("Municipality",            "text",     lambda a: a.municipality),
    ("Province",                "text",     lambda a: a.province),
    ("Latitude",                "float",    lambda a: a.lat),
    ("Longitude",               "float",    lambda a: a.lng),
]

HEADERS = [h for h, _k, _g in REGISTER_SPEC]

# Column widths copied from the client-formatted reference workbook (keyed by
# header so removing/reordering a column keeps the rest aligned).
_COL_WIDTHS = {
    "Luminaire Manufacturer": 26.0, "Luminaire Model": 21.3, "Luminaire Serial": 19.1,
    "Luminaire Year": 18.2, "Luminaire Wattage": 21.5, "Controller Manufacturer": 26.1,
    "Controller Model": 19.5, "Controller IMEI": 23.2, "Date of Installation": 21.5,
    "Work Order Number": 22.7, "Installer Type": 16.8, "Installer Name": 17.5,
    "Designation": 22.2, "Service Number": 18.9, "Company": 37.2, "Contractor Number": 21.9,
    "Region": 16.5, "Road Name": 14.3, "Suburb": 11.5, "Town": 18.0,
    "Municipality": 23.5, "Province": 18.0, "Latitude": 12.2, "Longitude": 13.8,
}

# Number formats per kind (text -> General). Matches the reference workbook.
_NUMFMT = {"int": "0", "float": "0.000000", "datetime": "yyyy\\-mm\\-dd\\ hh:mm", "imei": "@"}


def _csv_cell(kind, v):
    if v is None:
        return ""
    if kind == "datetime":
        return _iso(v) or ""
    return v  # int/float/str -> csv.writer str()s them


def build_register_csv(assets) -> str:
    """The 'standard' CSV (ISO-8601 dates, all values stringified)."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(HEADERS)
    for a in assets:
        w.writerow([_csv_cell(kind, get(a)) for _h, kind, get in REGISTER_SPEC])
    return buf.getvalue()


def _xlsx_cell(kind, v):
    if v is None:
        return None
    if kind == "imei":
        return str(v)  # keep verbatim; '@' format renders it as text
    if kind == "datetime" and getattr(v, "tzinfo", None) is not None:
        return v.replace(tzinfo=None)  # Excel has no tz; store naive wall-clock
    return v  # datetime(naive)/int/float/str native


def build_register_xlsx(assets) -> bytes:
    """The formatted Excel register: styled header, typed cells, frozen header,
    and an Excel Table (LuminaireRegister), matching the client reference file.
    openpyxl is imported lazily so the CSV path never depends on it."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    assets = list(assets)
    wb = Workbook()
    ws = wb.active
    ws.title = "nuevo_luminaire_register_" + datetime.now(timezone.utc).strftime("%Y%m")

    ws.append(HEADERS)
    hfill = PatternFill("solid", fgColor="FF1F4E78")
    hfont = Font(bold=True, color="FFFFFFFF", size=10)
    hcenter = Alignment(horizontal="center", vertical="center")
    for c in ws[1]:
        c.fill = hfill
        c.font = hfont
        c.alignment = hcenter

    for a in assets:
        ws.append([_xlsx_cell(kind, get(a)) for _h, kind, get in REGISTER_SPEC])

    last_row = ws.max_row
    for idx, (h, kind, _g) in enumerate(REGISTER_SPEC, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _COL_WIDTHS.get(h, 16)
        nf = _NUMFMT.get(kind)
        if nf:
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=idx).number_format = nf

    ws.freeze_panes = "A2"

    # An Excel Table needs at least one data row; skip it for an empty register.
    if assets:
        ref = "A1:%s%d" % (get_column_letter(len(HEADERS)), last_row)
        tbl = Table(displayName="LuminaireRegister", ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
            showFirstColumn=False, showLastColumn=False)
        ws.add_table(tbl)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
